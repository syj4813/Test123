# -*- coding: utf-8 -*-
"""
'한국철도공사_철도운행거리_전체' 원본은 시트별로 '계단형(staircase) 삼각행렬'
형태다. 각 행(row)에서 역명이 있는 셀의 열(col) 위치가 행이 진행될수록
1칸씩 이동(순증 또는 역순으로 -1)하며, 그 뒤(혹은 앞)로 이어지는 숫자들은
해당 역에서 그 열에 대응하는 후속 역까지의 '역간 실제 누적거리'다.

이 스크립트는 그 구조를 일반화해서 파싱, (역A, 역B, 인접거리km) 엣지 목록으로
변환한다. 인접한 역끼리의 거리만 뽑으면(전체 조합 아님) 그래프로 쌓아
다익스트라로 임의 두 역 사이 거리를 복원할 수 있다.
"""

import csv
import openpyxl

SRC = "한국철도공사_철도운행거리_전체_20240901.xlsx"


def is_number(v):
    return isinstance(v, (int, float))


def extract_chains(ws):
    """시트 하나에서 (역명, 행, 열) 후보를 모두 찾고,
    열이 +1씩 순증하거나 -1씩 순감하는 연속열을 하나의 '체인'으로 묶는다.
    반환: [ [(name, row, col), ...], ... ]  (체인 리스트)
    """
    candidates = []
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c, val in enumerate(row, start=1):
            if isinstance(val, str):
                v = val.strip()
                if not v or v in ("-",):
                    continue
                # 줄바꿈 포함된 라벨('연결선\n(경부선)' 등)이나 설명 텍스트는 역명이 아님
                if "\n" in v or "운행거리" in v or "연결선" in v or v.startswith("("):
                    continue
                candidates.append((v, r, c))

    candidates.sort(key=lambda x: x[1])  # row 순서

    chains = []
    open_chains = []  # 각 원소: {"dir": +1/-1, "last_row":..., "last_col":..., "items":[...]}

    for name, r, c in candidates:
        attached = False
        for ch in open_chains:
            if r == ch["last_row"] + 1 and c == ch["last_col"] + ch["dir"]:
                ch["items"].append((name, r, c))
                ch["last_row"], ch["last_col"] = r, c
                attached = True
                break
        if not attached:
            # 새 체인 시작 (방향은 다음 후보가 붙을 때 확정하되, 기본은 +1 시도)
            open_chains.append({"dir": 1, "last_row": r, "last_col": c, "items": [(name, r, c)]})

    for ch in open_chains:
        if len(ch["items"]) >= 2:
            chains.append(ch["items"])
    return chains


def chain_edges(ws, chain):
    """체인(역 순서 리스트) -> [(역A, 역B, km), ...] 인접거리 엣지"""
    edges = []
    grid_cache = {}

    def cell(r, c):
        if (r, c) not in grid_cache:
            grid_cache[(r, c)] = ws.cell(row=r, column=c).value
        return grid_cache[(r, c)]

    for i in range(len(chain) - 1):
        name_a, r_a, c_a = chain[i]
        name_b, r_b, c_b = chain[i + 1]
        # 상삼각형(정방향): A행, B열 위치에 값이 있음 (예: 경부KTX 시트)
        val = cell(r_a, c_b)
        # 하삼각형(역방향): B행, A열 위치에 값이 있음 (예: 영동/수인/서해 시트)
        if not (is_number(val) and val > 0):
            val = cell(r_b, c_a)
        if is_number(val) and val > 0:
            edges.append((name_a, name_b, round(float(val), 2)))
    return edges


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    all_edges = []
    summary = []

    # 수도권 광역전철 전용 시트는 제외한다.
    # 이유: (1) KTX/무궁화/새마을 중장거리 비교에는 기여가 없고,
    #       (2) '구룡'처럼 지방 노선(경전선)과 서울 지하철 역명이 동일해
    #           그래프를 잘못 이어버리는 동명이역 충돌 위험이 있음.
    EXCLUDE_SHEETS = {"000000", "17-과천,분당,일산", "19-수인", "20-서해"}

    for ws in wb.worksheets:
        if ws.title in EXCLUDE_SHEETS:
            continue
        chains = extract_chains(ws)
        sheet_edges = []
        for ch in chains:
            sheet_edges.extend(chain_edges(ws, ch))
        all_edges.extend(sheet_edges)
        summary.append((ws.title, len(chains), len(sheet_edges)))

    print(f"{'시트':25s} {'체인수':>6s} {'엣지수':>6s}")
    for title, nchain, nedge in summary:
        print(f"{title:25s} {nchain:>6d} {nedge:>6d}")
    print(f"\n총 엣지: {len(all_edges)}")

    # 역명 중복 엣지(다른 시트에서 같은 두 역이 다시 등장) -> 최솟값 유지, 값 충돌 로그
    best = {}
    conflicts = []
    for a, b, km in all_edges:
        key = tuple(sorted([a, b]))
        if key in best and abs(best[key] - km) > 0.5:
            conflicts.append((key, best[key], km))
        if key not in best or km < best[key]:
            best[key] = km

    print(f"고유 역쌍(엣지): {len(best)}, 값 충돌(0.5km초과 차이): {len(conflicts)}")
    if conflicts[:10]:
        print("충돌 예시:")
        for key, v1, v2 in conflicts[:10]:
            print(f"  {key}: {v1} vs {v2}")

    with open("rail_edges.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for (a, b), km in best.items():
            w.writerow([a, b, km])

    stations = set()
    for a, b in best.keys():
        stations.add(a)
        stations.add(b)
    print(f"\n고유 역명 수: {len(stations)}")


if __name__ == "__main__":
    main()
